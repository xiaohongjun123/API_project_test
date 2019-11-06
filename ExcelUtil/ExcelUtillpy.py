from openpyxl import load_workbook
class ReadExcel(object):
    def __init__(self,excelname,excelsheet):
        self.wb=load_workbook(excelname)
        self.sheet=self.wb.get_sheet_by_name(excelsheet)
        self.maxRownum=self.sheet.max_row#获取最大行数
        self.maxColumn=self.sheet.max_column#获取最大列数
        print(self.maxRownum,self.maxColumn)


    def getValue(self):
        datalist=[]
        for row in self.sheet.rows[1:self.maxRownum]:
            list=[]
            for column in range(3,self.maxColumn):
                list.append(row[column].value)
            datalist.append(list)
        return datalist

    def writeValue(self,results):
        for row in range(1,self.maxRownum):
            for column in range(1,self.maxColumn+1):
                if self.sheet.cell(row=row,column=column).value==None and self.sheet.cell(row=row,column=column+1).value==None:
                    self.sheet.cell(row=row,column=column).value=results
                self.wb.save(r"G:\API_project_test\ExcelFile\casedata1.xlsx")



if __name__=="__main__":
    value=ReadExcel(r"G:\API_project_test\ExcelFile\casedata.xlsx", "API_Add_Org_Data")
    value.writeValue("Pass")

